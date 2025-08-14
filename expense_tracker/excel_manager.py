from datetime import date
import os
import pandas as pd
from openpyxl import Workbook, worksheet, load_workbook
from pathlib import Path
from utils import RESULTS_DIR, logger
class ExcelManager:
    def __init__(self):
        #set folder path to save excel
        self.results_path = RESULTS_DIR        
        #initialize logging
        self.logger = logger

    def create_monthly_workbook(self, month_year: list):
        """ Create workbook and worksheets if file does not exist
        excel name convention is Expense_<mon>_<yy>.xlsx.
        month_year input will already concatenate the month and year
        Workbook covers week from 14th of the month prior to 13th of the
        current month, e.g. Expense_Jul_25.xlsx should contain receipts from
        Jun 14 to Jul 13
        """

        expense_mth = month_year[0]
        expense_year = month_year[1]

        #get excel filename
        excel_filename = f"Expense_{expense_mth}_{expense_year}.xlsx"
        excel_filepath = RESULTS_DIR / excel_filename
        
        #get week number
        start_date = date(expense_year, expense_mth, 14)

        if int(expense_mth) == 12:
            end_date = date(expense_year + 1, 01, 14)
        else:
            end_date = date(expense_year, int(expense_mth) + 1, 14)

        start_week = self.determine_week_num(start_date)
        end_week = self.determine_week_num(end_date)
        
        # check if excel file exists already 
        if excel_filepath.exists():
            self.logger.info(f"file {excel_filename} already exists")
            return None
        else:       
            # if not Create Excel file with weekly sheets
            wb = Workbook()

            # Create summary sheet
            wb.create_sheet(title="Summary")
            ws = wb["Summary"]
            wb.active = ws

            #Format summary sheet
            ws["A2"] = "Spending by Week"
            ws.insert_cols(idx=1, amount=2)
            
            #Headers for Weekly Spend
            ws["A3"] = "Week No."
            ws["B3"] = "Total Spend"
            
            ws["E2"] = "Spending by credit_type"
            ws.insert_cols(idx=5, amount=2)

            #Headers for Spend by Credit Type
            ws["E3"] = "Credit Type"
            ws["F3"] = "Total Spend"

            #Create sheets for each week
            for week_num in (start_week, end_week):
                sheet_name = f"week_{week_num}"
                wb.create_sheet(title = sheet_name)
                
                #make active sheet
                active_sheet = wb[sheet_name]
                wb.active = active_sheet
                #insert 5 columns
                active_sheet.insert_cols(idx=1, amount=5)

                headers = ['date', 'description', 'amount', 'category', 'card_type','spending_type']            
                for col_num, header in enumerate(headers, start=1):
                    active_sheet.cell(row=1,column = col_num, value = header)
            wb.save(excel_filepath)

            return wb

    def add_transaction (self, expense_workbook: Workbook, transaction_details: dict):
        """
        Method will add transaction based on transaction date to the workbook
        """
        # Determine week number (19th-13th cycle)
        week_no = self.determine_week_num(transaction_details['date'])
        sheet_name = f"week_{week_no}"

        # Add to appropriate worksheet
        wb = load_workbook(expense_workbook)
        sheet_update = wb[sheet_name]

        new_row = self.last_row_in_column(sheet_update, "A") + 1
        
        #insert data
        sheet_update[f"A{new_row}"] = transaction_details['date']
        sheet_update[f"B{new_row}"] = transaction_details['description']
        sheet_update[f"C{new_row}"] = transaction_details['amount']
        sheet_update[f"D{new_row}"] = "" #leave category blank
        sheet_update[f"E{new_row}"] = transaction_details['card_type']
        sheet_update[f"F{new_row}"] = "" #leave spending_type blank for now
        
        wb.save(expense_workbook)

    def determine_week_num(self, date_to_check: date):
        """ Calculate week based on date using ISO calendar"""
        #Get iso_calendar representation (year, week no, weekday)
        iso_calendar = date_to_check.isocalendar()

        #get week number
        week_number =iso_calendar[1]
        
        return week_number

    def last_row_in_column(self, ws: worksheet, col_letter: str):
        
        for row in range(ws.max_row, 0, -1):
            if ws[f"{col_letter}{row}"].value not in (None, ""):
                return row
            
        return 0